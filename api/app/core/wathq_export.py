"""
WATHQ data export utilities for XLS format with nested JSON handling.
"""

from typing import Any, Dict, List, Optional
from io import BytesIO
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class WathqExporter:
    """Export WATHQ API responses to Excel format with nested JSON handling."""

    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.row_num = 1

    def _flatten_dict(self, data: Dict[str, Any], parent_key: str = "", sep: str = "_") -> Dict[str, Any]:
        """
        Flatten nested dictionaries and lists into a single level.
        
        Args:
            data: Dictionary to flatten
            parent_key: Parent key for nested items
            sep: Separator for nested keys
            
        Returns:
            Flattened dictionary
        """
        items = []
        
        for k, v in data.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                if v and isinstance(v[0], dict):
                    # For lists of dicts, create separate entries
                    for i, item in enumerate(v):
                        items.extend(self._flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                else:
                    # For simple lists, join with comma
                    items.append((new_key, ", ".join(str(x) for x in v)))
            else:
                items.append((new_key, v))
        
        return dict(items)

    def _format_value(self, value: Any) -> str:
        """Format value for Excel cell."""
        if value is None:
            return ""
        elif isinstance(value, bool):
            return "Yes" if value else "No"
        elif isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, indent=2)
        elif isinstance(value, datetime):
            return value.isoformat()
        else:
            return str(value)

    def _apply_header_style(self, cell):
        """Apply header styling to cell."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def _apply_data_style(self, cell):
        """Apply data styling to cell."""
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def export_single_record(self, data: Dict[str, Any], sheet_name: str = "Data") -> BytesIO:
        """
        Export a single WATHQ record to Excel.
        
        Args:
            data: WATHQ API response data
            sheet_name: Name of the worksheet
            
        Returns:
            BytesIO object containing the Excel file
        """
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.row_num = 1

        # Flatten the data
        flattened = self._flatten_dict(data)

        # Write headers
        headers = list(flattened.keys())
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.row_num, column=col_num)
            cell.value = header
            self._apply_header_style(cell)

        # Write data
        self.row_num += 1
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.row_num, column=col_num)
            cell.value = self._format_value(flattened[header])
            self._apply_data_style(cell)

        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(header), 20)
            self.worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Return BytesIO object
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def export_multiple_records(
        self, 
        records: List[Dict[str, Any]], 
        sheet_name: str = "Data"
    ) -> BytesIO:
        """
        Export multiple WATHQ records to Excel.
        
        Args:
            records: List of WATHQ API response data
            sheet_name: Name of the worksheet
            
        Returns:
            BytesIO object containing the Excel file
        """
        if not records:
            raise ValueError("No records to export")

        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        self.row_num = 1

        # Flatten all records and collect all unique keys
        flattened_records = [self._flatten_dict(record) for record in records]
        all_keys = set()
        for record in flattened_records:
            all_keys.update(record.keys())
        
        headers = sorted(list(all_keys))

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.row_num, column=col_num)
            cell.value = header
            self._apply_header_style(cell)

        # Write data rows
        for record in flattened_records:
            self.row_num += 1
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=self.row_num, column=col_num)
                cell.value = self._format_value(record.get(header))
                self._apply_data_style(cell)

        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(header), 20)
            self.worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Return BytesIO object
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def export_with_nested_sheets(
        self, 
        data: Dict[str, Any], 
        main_sheet_name: str = "Main"
    ) -> BytesIO:
        """
        Export WATHQ data with nested objects in separate sheets.
        
        Args:
            data: WATHQ API response data
            main_sheet_name: Name of the main worksheet
            
        Returns:
            BytesIO object containing the Excel file
        """
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)

        # Main sheet with flattened data (excluding nested objects)
        main_sheet = self.workbook.create_sheet(main_sheet_name)
        self.worksheet = main_sheet
        self.row_num = 1

        # Separate nested objects
        flat_data = {}
        nested_data = {}

        for key, value in data.items():
            if isinstance(value, dict):
                nested_data[key] = value
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                nested_data[key] = value
            else:
                flat_data[key] = value

        # Write main data
        headers = list(flat_data.keys())
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.row_num, column=col_num)
            cell.value = header
            self._apply_header_style(cell)

        self.row_num += 1
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.row_num, column=col_num)
            cell.value = self._format_value(flat_data[header])
            self._apply_data_style(cell)

        # Create sheets for nested objects
        for nested_key, nested_value in nested_data.items():
            if isinstance(nested_value, dict):
                self._create_nested_sheet(nested_key, nested_value)
            elif isinstance(nested_value, list):
                self._create_nested_sheet(nested_key, nested_value)

        # Adjust column widths for main sheet
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(header), 20)
            self.worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Return BytesIO object
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def _create_nested_sheet(self, sheet_name: str, data: Any):
        """Create a sheet for nested data."""
        # Sanitize sheet name (Excel has 31 char limit)
        sheet_name = sheet_name[:31]
        
        sheet = self.workbook.create_sheet(sheet_name)
        self.worksheet = sheet
        self.row_num = 1

        if isinstance(data, dict):
            # Single nested object
            flattened = self._flatten_dict(data)
            headers = list(flattened.keys())
            
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=self.row_num, column=col_num)
                cell.value = header
                self._apply_header_style(cell)

            self.row_num += 1
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=self.row_num, column=col_num)
                cell.value = self._format_value(flattened[header])
                self._apply_data_style(cell)

        elif isinstance(data, list):
            # List of objects
            flattened_records = [self._flatten_dict(item) if isinstance(item, dict) else {"value": item} for item in data]
            all_keys = set()
            for record in flattened_records:
                all_keys.update(record.keys())
            
            headers = sorted(list(all_keys))

            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=self.row_num, column=col_num)
                cell.value = header
                self._apply_header_style(cell)

            for record in flattened_records:
                self.row_num += 1
                for col_num, header in enumerate(headers, 1):
                    cell = self.worksheet.cell(row=self.row_num, column=col_num)
                    cell.value = self._format_value(record.get(header))
                    self._apply_data_style(cell)

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            self.worksheet.column_dimensions[column_letter].width = 25


# Convenience functions
def export_wathq_to_xls(data: Dict[str, Any], filename: str = "wathq_export.xlsx") -> BytesIO:
    """Export single WATHQ record to XLS."""
    exporter = WathqExporter()
    return exporter.export_single_record(data)


def export_wathq_records_to_xls(
    records: List[Dict[str, Any]], 
    filename: str = "wathq_records.xlsx"
) -> BytesIO:
    """Export multiple WATHQ records to XLS."""
    exporter = WathqExporter()
    return exporter.export_multiple_records(records)


def export_wathq_with_nested_sheets(
    data: Dict[str, Any], 
    filename: str = "wathq_detailed.xlsx"
) -> BytesIO:
    """Export WATHQ data with nested objects in separate sheets."""
    exporter = WathqExporter()
    return exporter.export_with_nested_sheets(data)
