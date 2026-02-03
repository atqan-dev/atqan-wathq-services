"""
WATHQ PDF Export API endpoints
"""

from typing import Any, Dict, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session

from app import models
from app.api import deps
from app.core.config import settings
from app.core.wathq_utils import get_tenant_wathq_key_by_slug
from app.models.wathq_pdf_data import WathqPDFData
from app.services.wathq_pdf_service import pdf_service
from app.wathq.commercial_registration.client import WathqClient

router = APIRouter()


def get_wathq_client_for_user(
    service_slug: str,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> WathqClient:
    """Get Wathq client instance for current user"""

    if isinstance(current_user, models.ManagementUser):
        # Management user - use global API key
        return WathqClient(
            api_key=settings.WATHQ_API_KEY, db=db, tenant_id=None, user_id=None
        )
    else:
        # Tenant user - use tenant-specific API key
        api_key = get_tenant_wathq_key_by_slug(
            db=db, tenant_id=current_user.tenant_id, service_slug=service_slug
        )

        # Fallback to system API key if tenant doesn't have specific key
        if not api_key:
            api_key = settings.WATHQ_API_KEY

        return WathqClient(
            api_key=api_key,
            db=db,
            tenant_id=current_user.tenant_id,
            user_id=current_user.id,
        )


@router.get("/commercial-registration/{cr_id}/pdf")
async def export_commercial_registration_pdf(
    cr_id: str,
    language: str = Query("ar", regex="^(ar|en)$"),
    template: Optional[str] = Query(None, description="Custom template name"),
    include_full_info: bool = Query(True, description="Include full CR info"),
    include_owners: bool = Query(False, description="Include owners info"),
    include_managers: bool = Query(False, description="Include managers info"),
    include_branches: bool = Query(False, description="Include branches info"),
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Export Commercial Registration data as PDF
    cr_id can be either the database ID or the CR number
    """
    try:
        from app.models.wathq_commercial_registration import CommercialRegistration

        # Try to get CR record from database first
        # Check if cr_id is a numeric database ID
        cr_number = cr_id
        try:
            db_id = int(cr_id)
            # It's a database ID, look up the CR record
            cr_record = (
                db.query(CommercialRegistration)
                .filter(CommercialRegistration.id == db_id)
                .first()
            )

            if cr_record and cr_record.cr_number:
                cr_number = cr_record.cr_number
                print(
                    f"Found CR record with database ID {db_id}, using cr_number: {cr_number}"
                )
            else:
                print(f"No CR record found for database ID {db_id}, using as cr_number")
        except ValueError:
            # It's not a number, assume it's already a cr_number
            print(f"Using provided cr_id as cr_number: {cr_id}")

        client = get_wathq_client_for_user("commercial-registration", db, current_user)

        # Fetch CR data using the actual cr_number
        if include_full_info:
            cr_data = await client.get_full_info(cr_number, language)
        else:
            cr_data = await client.get_basic_info(cr_number, language)

        # Fetch additional data if requested
        additional_sections = []

        if include_owners:
            try:
                owners_data = await client.get_owners(cr_number, language)
                if owners_data.get("owners"):
                    owners_content = pdf_service.create_table_from_wathq_data(
                        {"owners": owners_data["owners"]},
                        {"headers": ["الاسم", "الجنسية", "رقم الهوية", "النسبة"]},
                    )
                    additional_sections.append(
                        {
                            "title": "المالكون",
                            "content": f"<table class='data-table'>{owners_content[1]}</table>",
                            "page_break": True,
                        }
                    )
            except Exception:
                pass  # Continue if owners data not available

        if include_managers:
            try:
                managers_data = await client.get_managers(cr_number, language)
                if managers_data.get("managers"):
                    managers_content = pdf_service.create_table_from_wathq_data(
                        {"managers": managers_data["managers"]},
                        {"headers": ["الاسم", "المنصب", "رقم الهوية"]},
                    )
                    additional_sections.append(
                        {
                            "title": "المديرون",
                            "content": f"<table class='data-table'>{managers_content[1]}</table>",
                            "page_break": True,
                        }
                    )
            except Exception:
                pass

        if include_branches:
            try:
                branches_data = await client.get_branches(cr_number, language)
                if branches_data.get("branches"):
                    branches_content = pdf_service.create_table_from_wathq_data(
                        {"branches": branches_data["branches"]},
                        {"headers": ["اسم الفرع", "العنوان", "النشاط"]},
                    )
                    additional_sections.append(
                        {
                            "title": "الفروع",
                            "content": f"<table class='data-table'>{branches_content[1]}</table>",
                            "page_break": True,
                        }
                    )
            except Exception:
                pass

        # Create PDF data model
        pdf_data = pdf_service.create_commercial_registration_pdf(
            cr_data=cr_data, additional_content=None
        )

        # Add additional sections if any
        if additional_sections:
            if not pdf_data.sections:
                pdf_data.sections = []
            pdf_data.sections.extend(
                [
                    {
                        "title": section["title"],
                        "content": section["content"],
                        "page_break": section.get("page_break", False),
                    }
                    for section in additional_sections
                ]
            )

        # Generate and return PDF
        filename = f"commercial_registration_{cr_id}.pdf"

        # For commercial registration, always use the dedicated template
        # Ignore generic template requests that aren't compatible
        if template and template in [
            "wathq_modern_template.html",
            "wathq-modern-template.html",
        ]:
            template_name = "commercial_registration_pdf.html"
        else:
            template_name = template or "commercial_registration_pdf.html"

        return pdf_service.generate_pdf_response(
            pdf_data, filename=filename, template_name=template_name
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate CR PDF: {str(e)}"
        )


@router.post("/custom-document/pdf")
async def create_custom_pdf(
    pdf_data: WathqPDFData,
    template: Optional[str] = Query(None, description="Custom template name"),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Create custom PDF document from provided data
    """
    try:
        filename = f"{pdf_data.document_title or 'document'}.pdf"
        return pdf_service.generate_pdf_response(
            pdf_data, filename=filename, template_name=template
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate custom PDF: {str(e)}"
        )


@router.get("/preview/commercial-registration/{cr_id}")
async def preview_commercial_registration_html(
    cr_id: str,
    language: str = Query("ar", regex="^(ar|en)$"),
    template: Optional[str] = Query(None, description="Custom template name"),
    include_full_info: bool = Query(True, description="Include full CR info"),
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Preview Commercial Registration HTML before PDF generation
    """
    try:
        client = get_wathq_client_for_user("commercial-registration", db, current_user)

        # Fetch CR data
        if include_full_info:
            cr_data = await client.get_full_info(cr_id, language)
        else:
            cr_data = await client.get_basic_info(cr_id, language)

        # Create PDF data model
        pdf_data = pdf_service.create_commercial_registration_pdf(cr_data)

        # Generate HTML preview
        # For commercial registration, always use the dedicated template
        if template and template in [
            "wathq_modern_template.html",
            "wathq-modern-template.html",
        ]:
            template_name = "commercial_registration_pdf.html"
        else:
            template_name = template or "commercial_registration_pdf.html"

        html_content = pdf_service.preview_html(pdf_data, template_name=template_name)
        return Response(content=html_content, media_type="text/html")

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate preview: {str(e)}"
        )


@router.post("/preview/custom-document")
async def preview_custom_document_html(
    pdf_data: WathqPDFData,
    template: Optional[str] = Query(None, description="Custom template name"),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Preview custom document HTML before PDF generation
    """
    try:
        html_content = pdf_service.preview_html(pdf_data, template_name=template)
        return Response(content=html_content, media_type="text/html")
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate preview: {str(e)}"
        )


@router.post("/wathq-data/{service_type}/{identifier}/pdf")
async def export_wathq_data_pdf(
    service_type: str,
    identifier: str,
    language: str = Query("ar", regex="^(ar|en)$"),
    template: Optional[str] = Query(None, description="Custom template name"),
    custom_title: Optional[str] = Query(None, description="Custom document title"),
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Generic endpoint to export any WATHQ service data as PDF

    Supported service types:
    - commercial-registration
    - real-estate
    - employee
    - attorney
    - company-contract
    """
    try:
        # This would need to be expanded based on available WATHQ services
        if service_type == "commercial-registration":
            return await export_commercial_registration_pdf(
                cr_id=identifier,
                language=language,
                template=template,
                db=db,
                current_user=current_user,
            )
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Service type '{service_type}' not supported for PDF export",
            )

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to export {service_type} PDF: {str(e)}"
        )


@router.get("/database/commercial-registration/{cr_id}/pdf")
async def export_database_cr_pdf(
    cr_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Export Commercial Registration PDF from database record
    Uses stored CR data instead of fetching from Wathq API
    """
    try:
        from app.models.wathq_commercial_registration import CommercialRegistration
        from jinja2 import Template
        from datetime import datetime
        import pdfkit

        # Fetch CR from database with all relationships
        cr = (
            db.query(CommercialRegistration)
            .filter(CommercialRegistration.id == cr_id)
            .first()
        )

        if not cr:
            raise HTTPException(
                status_code=404,
                detail=f"Commercial Registration with ID {cr_id} not found",
            )

        # Load template - use commercial registration template with database mapping
        template_path = pdf_service.templates_dir / "cr_database_template_v2.html"
        if not template_path.exists():
            raise HTTPException(
                status_code=500, detail="Database CR template not found"
            )

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        template = Template(template_content)

        # Map database fields to template variables
        from datetime import datetime as dt

        # Create simple namespace objects for nested data
        class SimpleNamespace:
            def __init__(self, **kwargs):
                self.__dict__.update(kwargs)

        # Prepare entity_type object
        entity_type = SimpleNamespace(
            id=cr.entity_type_id,
            name=cr.entity_type_name,
            form_name=cr.entity_form_name,
        )

        # Prepare status object
        status_obj = None
        if cr.status_name:
            confirmation_date = None
            if cr.confirmation_date_gregorian:
                confirmation_date = SimpleNamespace(
                    gregorian=cr.confirmation_date_gregorian.strftime("%Y-%m-%d"),
                    hijri=cr.confirmation_date_hijri,
                )
            status_obj = SimpleNamespace(
                name=cr.status_name, confirmation_date=confirmation_date
            )

        # Prepare issue_date object
        issue_date = None
        if cr.issue_date_gregorian:
            issue_date = SimpleNamespace(
                gregorian=cr.issue_date_gregorian.strftime("%Y-%m-%d"),
                hijri=cr.issue_date_hijri,
            )

        # Prepare contact_info object
        contact_info = SimpleNamespace(
            phone=cr.contact_phone,
            mobile=cr.contact_mobile,
            email=cr.contact_email,
            website_url=cr.contact_website,
        )

        # Prepare capital object
        capital = None
        if cr.capital_info or cr.cr_capital:
            capital = SimpleNamespace(
                currency_name=(
                    cr.capital_info.currency_name if cr.capital_info else None
                ),
                capital=cr.cr_capital,
            )

        # Prepare activities list
        activities = []
        if cr.activities:
            for a in cr.activities:
                activities.append(
                    SimpleNamespace(id=a.activity_id, name=a.activity_name)
                )

        # Prepare parties list
        parties = []
        if cr.parties:
            for p in cr.parties:
                # Create nested identity object as expected by template
                identity = (
                    SimpleNamespace(id=p.identity_id, typeName=p.identity_type_name)
                    if p.identity_id
                    else None
                )

                # Create nationality object (template expects party.nationality.name)
                # We don't have nationality in parties table, so use None
                nationality = None

                parties.append(
                    SimpleNamespace(
                        name=p.name,
                        type_name=p.type_name,
                        identity=identity,
                        nationality=nationality,
                        partnership=[],  # Empty partnership list
                    )
                )

        # Prepare managers list
        managers = []
        if cr.managers:
            for m in cr.managers:
                managers.append(
                    SimpleNamespace(
                        name=m.name,
                        position=m.type_name,  # Map type_name to position
                        identity=m.identity_id,  # Template expects flat identity for managers
                        nationality_name=m.nationality_name,
                    )
                )

        # Prepare management structure object
        management = SimpleNamespace(
            structureName=cr.mgmt_structure_name,
            structureId=cr.mgmt_structure_id,
            managers=managers,  # Use the same managers list
        )

        template_data = {
            "document_title": f"السجل التجاري - {cr.cr_number}",
            "search_date": (
                cr.fetched_at.strftime("%Y-%m-%d")
                if cr.fetched_at
                else dt.now().strftime("%Y-%m-%d")
            ),
            "print_date": dt.now().strftime("%Y-%m-%d"),
            "name": cr.name,
            "name_lang_desc": cr.name_lang_desc,
            "cr_number": cr.cr_number,
            "cr_national_number": cr.cr_national_number,
            "cr_capital": cr.cr_capital,
            "company_duration": cr.company_duration,
            "version_no": cr.version_no,
            "entity_type": entity_type,
            "status": status_obj,
            "issue_date": issue_date,
            "headquarter_city": cr.headquarter_city_name,
            "contact_info": contact_info,
            "capital": capital,
            "activities": activities,
            "parties": parties,
            "managers": managers,
            "partners_nationality_name": cr.partners_nationality_name,
            "management": management,
        }

        html_content = template.render(**template_data)

        # Generate PDF
        pdf_options = {
            "page-size": "A4",
            "margin-top": "0mm",
            "margin-right": "0mm",
            "margin-bottom": "0mm",
            "margin-left": "0mm",
            "encoding": "UTF-8",
            "enable-local-file-access": None,
            "print-media-type": None,
            "orientation": "portrait",
        }

        pdf_bytes = pdfkit.from_string(html_content, False, options=pdf_options)

        filename = f"commercial_registration_{cr.cr_number}_{cr_id}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate database CR PDF: {str(e)}"
        )


@router.get("/database/commercial-registration/{cr_id}/preview")
async def preview_database_cr_html(
    cr_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Preview Commercial Registration HTML from database record
    """
    try:
        from app.models.wathq_commercial_registration import CommercialRegistration
        from jinja2 import Template
        from datetime import datetime as dt

        # Fetch CR from database with all relationships
        cr = (
            db.query(CommercialRegistration)
            .filter(CommercialRegistration.id == cr_id)
            .first()
        )

        if not cr:
            raise HTTPException(
                status_code=404,
                detail=f"Commercial Registration with ID {cr_id} not found",
            )

        # Load template
        template_path = pdf_service.templates_dir / "cr_database_template_v2.html"
        if not template_path.exists():
            raise HTTPException(
                status_code=500, detail="Database CR template not found"
            )

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        template = Template(template_content)

        # Create simple namespace objects for nested data (same as PDF endpoint)
        class SimpleNamespace:
            def __init__(self, **kwargs):
                self.__dict__.update(kwargs)

        entity_type = SimpleNamespace(
            id=cr.entity_type_id,
            name=cr.entity_type_name,
            form_name=cr.entity_form_name,
        )

        status_obj = None
        if cr.status_name:
            confirmation_date = None
            if cr.confirmation_date_gregorian:
                confirmation_date = SimpleNamespace(
                    gregorian=cr.confirmation_date_gregorian.strftime("%Y-%m-%d"),
                    hijri=cr.confirmation_date_hijri,
                )
            status_obj = SimpleNamespace(
                name=cr.status_name, confirmation_date=confirmation_date
            )

        issue_date = None
        if cr.issue_date_gregorian:
            issue_date = SimpleNamespace(
                gregorian=cr.issue_date_gregorian.strftime("%Y-%m-%d"),
                hijri=cr.issue_date_hijri,
            )

        contact_info = SimpleNamespace(
            phone=cr.contact_phone,
            mobile=cr.contact_mobile,
            email=cr.contact_email,
            website_url=cr.contact_website,
        )

        capital = None
        if cr.capital_info or cr.cr_capital:
            capital = SimpleNamespace(
                currency_name=(
                    cr.capital_info.currency_name if cr.capital_info else None
                ),
                capital=cr.cr_capital,
            )

        activities = (
            [
                SimpleNamespace(id=a.activity_id, name=a.activity_name)
                for a in cr.activities
            ]
            if cr.activities
            else []
        )

        parties = []
        if cr.parties:
            for p in cr.parties:
                identity = (
                    SimpleNamespace(id=p.identity_id, typeName=p.identity_type_name)
                    if p.identity_id
                    else None
                )
                parties.append(
                    SimpleNamespace(
                        name=p.name,
                        type_name=p.type_name,
                        identity=identity,
                        nationality=None,
                        partnership=[],
                    )
                )

        managers = []
        if cr.managers:
            for m in cr.managers:
                managers.append(
                    SimpleNamespace(
                        name=m.name,
                        position=m.type_name,
                        identity=m.identity_id,
                        nationality_name=m.nationality_name,
                    )
                )

        # Prepare management structure object
        management = SimpleNamespace(
            structureName=cr.mgmt_structure_name,
            structureId=cr.mgmt_structure_id,
            managers=managers,
        )

        template_data = {
            "document_title": f"السجل التجاري - {cr.cr_number}",
            "search_date": (
                cr.fetched_at.strftime("%Y-%m-%d")
                if cr.fetched_at
                else dt.now().strftime("%Y-%m-%d")
            ),
            "print_date": dt.now().strftime("%Y-%m-%d"),
            "name": cr.name,
            "name_lang_desc": cr.name_lang_desc,
            "cr_number": cr.cr_number,
            "cr_national_number": cr.cr_national_number,
            "cr_capital": cr.cr_capital,
            "company_duration": cr.company_duration,
            "version_no": cr.version_no,
            "entity_type": entity_type,
            "status": status_obj,
            "issue_date": issue_date,
            "headquarter_city": cr.headquarter_city_name,
            "contact_info": contact_info,
            "capital": capital,
            "activities": activities,
            "parties": parties,
            "managers": managers,
            "partners_nationality_name": cr.partners_nationality_name,
            "management": management,
        }

        html_content = template.render(**template_data)

        return Response(content=html_content, media_type="text/html")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate database CR preview: {str(e)}"
        )


@router.get("/database/corporate-contract/{contract_id}/pdf")
async def export_database_corporate_contract_pdf(
    contract_id: int,
    db: Session = Depends(deps.get_db),
) -> Response:
    """
    Export Corporate Contract PDF from database record
    Uses stored contract data instead of fetching from Wathq API
    """
    try:
        from app.models.wathq_corporate_contract import CorporateContract
        from jinja2 import Template
        from datetime import datetime as dt
        from sqlalchemy.orm import joinedload
        import pdfkit

        # Fetch contract from database with all relationships using eager loading
        contract = (
            db.query(CorporateContract)
            .options(
                joinedload(CorporateContract.stocks),
                joinedload(CorporateContract.parties),
                joinedload(CorporateContract.managers),
                joinedload(CorporateContract.management_config),
                joinedload(CorporateContract.activities),
                joinedload(CorporateContract.articles),
                joinedload(CorporateContract.decisions),
                joinedload(CorporateContract.notification_channels),
            )
            .filter(CorporateContract.id == contract_id)
            .first()
        )

        if not contract:
            raise HTTPException(
                status_code=404,
                detail=f"Corporate Contract with ID {contract_id} not found",
            )

        # Load template
        template_path = (
            pdf_service.templates_dir / "corporate_contracts_database_template.html"
        )
        if not template_path.exists():
            raise HTTPException(
                status_code=500, detail="Corporate Contract template not found"
            )

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        template = Template(template_content)

        # Prepare template data
        template_data = {
            "document_title": f"عقد تأسيس الشركة - {contract.entity_name or contract.cr_number}",
            "contract_date": (
                contract.contract_date.strftime("%Y-%m-%d")
                if contract.contract_date
                else "غير محدد"
            ),
            "print_date": dt.now().strftime("%Y-%m-%d"),
            "contract_id": contract.contract_id,
            "contract_copy_number": contract.contract_copy_number,
            "cr_national_number": contract.cr_national_number,
            "cr_number": contract.cr_number,
            "entity_name": contract.entity_name,
            "entity_name_lang_desc": contract.entity_name_lang_desc,
            "company_duration": contract.company_duration,
            "headquarter_city_name": contract.headquarter_city_name,
            "is_license_based": contract.is_license_based,
            "entity_type_name": contract.entity_type_name,
            "entity_form_name": contract.entity_form_name,
            "fiscal_calendar_type": contract.fiscal_calendar_type,
            "fiscal_year_end_month": contract.fiscal_year_end_month,
            "fiscal_year_end_day": contract.fiscal_year_end_day,
            "fiscal_year_end_year": contract.fiscal_year_end_year,
            "currency_name": contract.currency_name,
            "total_capital": contract.total_capital,
            "paid_capital": contract.paid_capital,
            "cash_capital": contract.cash_capital,
            "in_kind_capital": contract.in_kind_capital,
            "is_set_aside_enabled": contract.is_set_aside_enabled,
            "profit_allocation_percentage": contract.profit_allocation_percentage,
            "profit_allocation_purpose": contract.profit_allocation_purpose,
            "additional_decision_text": contract.additional_decision_text,
            "stocks": (
                [
                    {
                        "stock_type_name": getattr(s, "stock_type_name", None),
                        "stock_count": getattr(s, "stock_count", None),
                        "stock_value": getattr(s, "stock_value", None),
                    }
                    for s in (contract.stocks or [])
                ]
                if hasattr(contract, "stocks")
                else []
            ),
            "parties": (
                [
                    {
                        "name": getattr(p, "name", None),
                        "type_name": getattr(p, "type_name", None),
                        "identity_number": getattr(p, "identity_number", None),
                        "identity_type": getattr(p, "identity_type", None),
                        "nationality": getattr(p, "nationality", None),
                    }
                    for p in (contract.parties or [])
                ]
                if hasattr(contract, "parties")
                else []
            ),
            "managers": (
                [
                    {
                        "name": getattr(m, "name", None),
                        "type_name": getattr(m, "type_name", None),
                        "is_licensed": getattr(m, "is_licensed", None),
                        "identity_number": getattr(m, "identity_number", None),
                        "nationality": getattr(m, "nationality", None),
                        "position_name": getattr(m, "position_name", None),
                    }
                    for m in (contract.managers or [])
                ]
                if hasattr(contract, "managers")
                else []
            ),
            "management_config": (
                {
                    "structure_name": getattr(
                        contract.management_config, "structure_name", None
                    ),
                    "meeting_quorum_name": getattr(
                        contract.management_config, "meeting_quorum_name", None
                    ),
                    "can_delegate_attendance": getattr(
                        contract.management_config, "can_delegate_attendance", None
                    ),
                    "term_years": getattr(
                        contract.management_config, "term_years", None
                    ),
                }
                if hasattr(contract, "management_config") and contract.management_config
                else None
            ),
            "activities": (
                [
                    {
                        "activity_id": getattr(a, "activity_id", None),
                        "activity_name": getattr(a, "activity_name", None),
                    }
                    for a in (contract.activities or [])
                ]
                if hasattr(contract, "activities")
                else []
            ),
            "articles": (
                [
                    {
                        "original_id": getattr(art, "original_id", None),
                        "article_text": getattr(art, "article_text", None),
                        "part_name": getattr(art, "part_name", None),
                    }
                    for art in (contract.articles or [])
                ]
                if hasattr(contract, "articles")
                else []
            ),
            "decisions": (
                [
                    {
                        "decision_name": getattr(d, "decision_name", None),
                        "approve_percentage": getattr(d, "approve_percentage", None),
                    }
                    for d in (contract.decisions or [])
                ]
                if hasattr(contract, "decisions")
                else []
            ),
            "notification_channels": (
                [
                    {"channel_name": getattr(nc, "channel_name", None)}
                    for nc in (contract.notification_channels or [])
                ]
                if hasattr(contract, "notification_channels")
                else []
            ),
            "contact_info": {
                "phone_no": None,
                "mobile_no": None,
                "email": None,
                "website_url": None,
            },
        }

        html_content = template.render(**template_data)

        # Generate PDF
        pdf_options = {
            "page-size": "A4",
            "margin-top": "0mm",
            "margin-right": "0mm",
            "margin-bottom": "0mm",
            "margin-left": "0mm",
            "encoding": "UTF-8",
            "enable-local-file-access": None,
            "print-media-type": None,
            "orientation": "portrait",
        }

        pdf_bytes = pdfkit.from_string(html_content, False, options=pdf_options)

        filename = f"corporate_contract_{contract.cr_number}_{contract_id}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate corporate contract PDF: {str(e)}",
        )


@router.get("/database/corporate-contract/{contract_id}/preview")
async def preview_database_corporate_contract_html(
    contract_id: int,
    db: Session = Depends(deps.get_db),
) -> Response:
    """
    Preview Corporate Contract HTML from database record
    """
    try:
        from app.models.wathq_corporate_contract import CorporateContract
        from jinja2 import Template
        from datetime import datetime as dt
        from sqlalchemy.orm import joinedload

        # Fetch contract from database with all relationships using eager loading
        contract = (
            db.query(CorporateContract)
            .options(
                joinedload(CorporateContract.stocks),
                joinedload(CorporateContract.parties),
                joinedload(CorporateContract.managers),
                joinedload(CorporateContract.management_config),
                joinedload(CorporateContract.activities),
                joinedload(CorporateContract.articles),
                joinedload(CorporateContract.decisions),
                joinedload(CorporateContract.notification_channels),
            )
            .filter(CorporateContract.id == contract_id)
            .first()
        )

        if not contract:
            raise HTTPException(
                status_code=404,
                detail=f"Corporate Contract with ID {contract_id} not found",
            )

        # Load template
        template_path = (
            pdf_service.templates_dir / "corporate_contracts_database_template.html"
        )
        if not template_path.exists():
            raise HTTPException(
                status_code=500, detail="Corporate Contract template not found"
            )

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        template = Template(template_content)

        # Prepare template data (same as PDF endpoint)
        template_data = {
            "document_title": f"عقد تأسيس الشركة - {contract.entity_name or contract.cr_number}",
            "contract_date": (
                contract.contract_date.strftime("%Y-%m-%d")
                if contract.contract_date
                else "غير محدد"
            ),
            "print_date": dt.now().strftime("%Y-%m-%d"),
            "contract_id": contract.contract_id,
            "contract_copy_number": contract.contract_copy_number,
            "cr_national_number": contract.cr_national_number,
            "cr_number": contract.cr_number,
            "entity_name": contract.entity_name,
            "entity_name_lang_desc": contract.entity_name_lang_desc,
            "company_duration": contract.company_duration,
            "headquarter_city_name": contract.headquarter_city_name,
            "is_license_based": contract.is_license_based,
            "entity_type_name": contract.entity_type_name,
            "entity_form_name": contract.entity_form_name,
            "fiscal_calendar_type": contract.fiscal_calendar_type,
            "fiscal_year_end_month": contract.fiscal_year_end_month,
            "fiscal_year_end_day": contract.fiscal_year_end_day,
            "fiscal_year_end_year": contract.fiscal_year_end_year,
            "currency_name": contract.currency_name,
            "total_capital": contract.total_capital,
            "paid_capital": contract.paid_capital,
            "cash_capital": contract.cash_capital,
            "in_kind_capital": contract.in_kind_capital,
            "is_set_aside_enabled": contract.is_set_aside_enabled,
            "profit_allocation_percentage": contract.profit_allocation_percentage,
            "profit_allocation_purpose": contract.profit_allocation_purpose,
            "additional_decision_text": contract.additional_decision_text,
            "stocks": (
                [
                    {
                        "stock_type_name": getattr(s, "stock_type_name", None),
                        "stock_count": getattr(s, "stock_count", None),
                        "stock_value": getattr(s, "stock_value", None),
                    }
                    for s in (contract.stocks or [])
                ]
                if hasattr(contract, "stocks")
                else []
            ),
            "parties": (
                [
                    {
                        "name": getattr(p, "name", None),
                        "type_name": getattr(p, "type_name", None),
                        "identity_number": getattr(p, "identity_number", None),
                        "identity_type": getattr(p, "identity_type", None),
                        "nationality": getattr(p, "nationality", None),
                    }
                    for p in (contract.parties or [])
                ]
                if hasattr(contract, "parties")
                else []
            ),
            "managers": (
                [
                    {
                        "name": getattr(m, "name", None),
                        "type_name": getattr(m, "type_name", None),
                        "is_licensed": getattr(m, "is_licensed", None),
                        "identity_number": getattr(m, "identity_number", None),
                        "nationality": getattr(m, "nationality", None),
                        "position_name": getattr(m, "position_name", None),
                    }
                    for m in (contract.managers or [])
                ]
                if hasattr(contract, "managers")
                else []
            ),
            "management_config": (
                {
                    "structure_name": getattr(
                        contract.management_config, "structure_name", None
                    ),
                    "meeting_quorum_name": getattr(
                        contract.management_config, "meeting_quorum_name", None
                    ),
                    "can_delegate_attendance": getattr(
                        contract.management_config, "can_delegate_attendance", None
                    ),
                    "term_years": getattr(
                        contract.management_config, "term_years", None
                    ),
                }
                if hasattr(contract, "management_config") and contract.management_config
                else None
            ),
            "activities": (
                [
                    {
                        "activity_id": getattr(a, "activity_id", None),
                        "activity_name": getattr(a, "activity_name", None),
                    }
                    for a in (contract.activities or [])
                ]
                if hasattr(contract, "activities")
                else []
            ),
            "articles": (
                [
                    {
                        "original_id": getattr(art, "original_id", None),
                        "article_text": getattr(art, "article_text", None),
                        "part_name": getattr(art, "part_name", None),
                    }
                    for art in (contract.articles or [])
                ]
                if hasattr(contract, "articles")
                else []
            ),
            "decisions": (
                [
                    {
                        "decision_name": getattr(d, "decision_name", None),
                        "approve_percentage": getattr(d, "approve_percentage", None),
                    }
                    for d in (contract.decisions or [])
                ]
                if hasattr(contract, "decisions")
                else []
            ),
            "notification_channels": (
                [
                    {"channel_name": getattr(nc, "channel_name", None)}
                    for nc in (contract.notification_channels or [])
                ]
                if hasattr(contract, "notification_channels")
                else []
            ),
            "contact_info": {
                "phone_no": None,
                "mobile_no": None,
                "email": None,
                "website_url": None,
            },
        }

        # Add JavaScript for export functionality
        export_script = f"""
        <script>
            function exportToPDF() {{
                window.location.href = '/api/v1/wathq/pdf/database/corporate-contract/{contract_id}/pdf';
            }}
        </script>
        """

        html_content = template.render(**template_data)
        html_content = html_content.replace("</body>", f"{export_script}</body>")

        return Response(content=html_content, media_type="text/html")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate corporate contract preview: {str(e)}",
        )


@router.get("/templates")
async def list_available_templates(
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Dict[str, Any]:
    """
    List available PDF templates
    """
    try:
        templates_dir = pdf_service.templates_dir
        templates = []

        if templates_dir.exists():
            for template_file in templates_dir.glob("*.html"):
                templates.append(
                    {
                        "name": template_file.stem,
                        "filename": template_file.name,
                        "path": str(template_file),
                    }
                )

        return {
            "templates": templates,
            "default_template": pdf_service.default_template,
        }

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to list templates: {str(e)}"
        )


@router.get("/database/employee/{employee_id}/pdf")
async def export_database_employee_pdf(
    employee_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Export Employee PDF from database record
    Uses stored employee data from wathq schema
    """
    try:
        from app.models.wathq_employee import Employee, EmploymentDetail
        from jinja2 import Template
        from datetime import datetime as dt
        from sqlalchemy.orm import joinedload
        import pdfkit

        # Fetch employee from database with employment details
        employee = (
            db.query(Employee)
            .options(joinedload(Employee.employment_details))
            .filter(Employee.employee_id == employee_id)
            .first()
        )

        if not employee:
            raise HTTPException(
                status_code=404, detail=f"Employee with ID {employee_id} not found"
            )

        # Load template
        template_path = pdf_service.templates_dir / "employee_database_template_v2.html"
        if not template_path.exists():
            raise HTTPException(status_code=500, detail="Employee template not found")

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        template = Template(template_content)

        # Prepare employment details
        employment_details = []
        if employee.employment_details:
            for detail in employee.employment_details:
                employment_details.append(
                    {
                        "employer": detail.employer,
                        "status": detail.status,
                        "basic_wage": (
                            f"{detail.basic_wage:.2f}" if detail.basic_wage else "0.00"
                        ),
                        "housing_allowance": (
                            f"{detail.housing_allowance:.2f}"
                            if detail.housing_allowance
                            else "0.00"
                        ),
                        "other_allowance": (
                            f"{detail.other_allowance:.2f}"
                            if detail.other_allowance
                            else "0.00"
                        ),
                        "full_wage": (
                            f"{detail.full_wage:.2f}" if detail.full_wage else "0.00"
                        ),
                    }
                )

        template_data = {
            "document_title": f"بيانات الموظف - {employee.name or employee_id}",
            "search_date": (
                employee.fetched_at.strftime("%Y-%m-%d")
                if employee.fetched_at
                else dt.now().strftime("%Y-%m-%d")
            ),
            "print_date": dt.now().strftime("%Y-%m-%d"),
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "nationality": employee.nationality,
            "working_months": employee.working_months,
            "fetched_at": (
                employee.fetched_at.strftime("%Y-%m-%d %H:%M:%S")
                if employee.fetched_at
                else "غير محدد"
            ),
            "employment_details": employment_details,
        }

        html_content = template.render(**template_data)

        # Generate PDF
        pdf_options = {
            "page-size": "A4",
            "margin-top": "0mm",
            "margin-right": "0mm",
            "margin-bottom": "0mm",
            "margin-left": "0mm",
            "encoding": "UTF-8",
            "enable-local-file-access": None,
            "print-media-type": None,
            "orientation": "portrait",
        }

        pdf_bytes = pdfkit.from_string(html_content, False, options=pdf_options)

        # Use only employee_id in filename to avoid encoding issues
        from urllib.parse import quote

        filename = f"employee_{employee_id}.pdf"
        encoded_filename = quote(filename)

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate employee PDF: {str(e)}"
        )


@router.get("/database/employee/{employee_id}/preview")
async def preview_database_employee_html(
    employee_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Preview Employee HTML from database record
    """
    try:
        from app.models.wathq_employee import Employee, EmploymentDetail
        from jinja2 import Template
        from datetime import datetime as dt
        from sqlalchemy.orm import joinedload

        # Fetch employee from database with employment details
        employee = (
            db.query(Employee)
            .options(joinedload(Employee.employment_details))
            .filter(Employee.employee_id == employee_id)
            .first()
        )

        if not employee:
            raise HTTPException(
                status_code=404, detail=f"Employee with ID {employee_id} not found"
            )

        # Load template
        template_path = pdf_service.templates_dir / "employee_database_template_v2.html"
        if not template_path.exists():
            raise HTTPException(status_code=500, detail="Employee template not found")

        with open(template_path, "r", encoding="utf-8") as f:
            template_content = f.read()

        template = Template(template_content)

        # Prepare employment details
        employment_details = []
        if employee.employment_details:
            for detail in employee.employment_details:
                employment_details.append(
                    {
                        "employer": detail.employer,
                        "status": detail.status,
                        "basic_wage": (
                            f"{detail.basic_wage:.2f}" if detail.basic_wage else "0.00"
                        ),
                        "housing_allowance": (
                            f"{detail.housing_allowance:.2f}"
                            if detail.housing_allowance
                            else "0.00"
                        ),
                        "other_allowance": (
                            f"{detail.other_allowance:.2f}"
                            if detail.other_allowance
                            else "0.00"
                        ),
                        "full_wage": (
                            f"{detail.full_wage:.2f}" if detail.full_wage else "0.00"
                        ),
                    }
                )

        template_data = {
            "document_title": f"بيانات الموظف - {employee.name or employee_id}",
            "search_date": (
                employee.fetched_at.strftime("%Y-%m-%d")
                if employee.fetched_at
                else dt.now().strftime("%Y-%m-%d")
            ),
            "print_date": dt.now().strftime("%Y-%m-%d"),
            "employee_id": employee.employee_id,
            "employee_name": employee.name,
            "nationality": employee.nationality,
            "working_months": employee.working_months,
            "fetched_at": (
                employee.fetched_at.strftime("%Y-%m-%d %H:%M:%S")
                if employee.fetched_at
                else "غير محدد"
            ),
            "employment_details": employment_details,
        }

        # Add JavaScript for export functionality
        export_script = f"""
        <script>
            function exportToPDF() {{
                window.location.href = '/api/v1/wathq/pdf/database/employee/{employee_id}/pdf';
            }}
            
            function exportToJSON() {{
                window.location.href = '/api/v1/wathq/pdf/database/employee/{employee_id}/json';
            }}
            
            function exportToCSV() {{
                window.location.href = '/api/v1/wathq/pdf/database/employee/{employee_id}/csv';
            }}
            
            function exportToExcel() {{
                window.location.href = '/api/v1/wathq/pdf/database/employee/{employee_id}/excel';
            }}
        </script>
        """

        html_content = template.render(**template_data)
        html_content = html_content.replace("</body>", f"{export_script}</body>")

        return Response(content=html_content, media_type="text/html")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to generate employee preview: {str(e)}"
        )


@router.get("/database/employee/{employee_id}/json")
async def export_database_employee_json(
    employee_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Export Employee data as JSON
    """
    try:
        from app.models.wathq_employee import Employee
        from sqlalchemy.orm import joinedload
        from urllib.parse import quote
        import json

        # Fetch employee from database with employment details
        employee = (
            db.query(Employee)
            .options(joinedload(Employee.employment_details))
            .filter(Employee.employee_id == employee_id)
            .first()
        )

        if not employee:
            raise HTTPException(
                status_code=404, detail=f"Employee with ID {employee_id} not found"
            )

        # Prepare data
        employee_data = {
            "employee_id": employee.employee_id,
            "name": employee.name,
            "nationality": employee.nationality,
            "working_months": employee.working_months,
            "fetched_at": (
                employee.fetched_at.isoformat() if employee.fetched_at else None
            ),
            "created_at": (
                employee.created_at.isoformat() if employee.created_at else None
            ),
            "updated_at": (
                employee.updated_at.isoformat() if employee.updated_at else None
            ),
            "employment_details": [],
        }

        if employee.employment_details:
            for detail in employee.employment_details:
                employee_data["employment_details"].append(
                    {
                        "employment_id": detail.employment_id,
                        "employer": detail.employer,
                        "status": detail.status,
                        "basic_wage": (
                            float(detail.basic_wage) if detail.basic_wage else 0.0
                        ),
                        "housing_allowance": (
                            float(detail.housing_allowance)
                            if detail.housing_allowance
                            else 0.0
                        ),
                        "other_allowance": (
                            float(detail.other_allowance)
                            if detail.other_allowance
                            else 0.0
                        ),
                        "full_wage": (
                            float(detail.full_wage) if detail.full_wage else 0.0
                        ),
                    }
                )

        json_content = json.dumps(employee_data, ensure_ascii=False, indent=2)

        filename = f"employee_{employee_id}.json"
        encoded_filename = quote(filename)

        return Response(
            content=json_content,
            media_type="application/json; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to export employee JSON: {str(e)}"
        )


@router.get("/database/employee/{employee_id}/csv")
async def export_database_employee_csv(
    employee_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Export Employee data as CSV with UTF-8 encoding for Arabic support
    """
    try:
        from app.models.wathq_employee import Employee
        from sqlalchemy.orm import joinedload
        from urllib.parse import quote
        import csv
        from io import StringIO, BytesIO

        # Fetch employee from database with employment details
        employee = (
            db.query(Employee)
            .options(joinedload(Employee.employment_details))
            .filter(Employee.employee_id == employee_id)
            .first()
        )

        if not employee:
            raise HTTPException(
                status_code=404, detail=f"Employee with ID {employee_id} not found"
            )

        # Create CSV with UTF-8 encoding
        output = StringIO()
        writer = csv.writer(output)

        # Write employee basic info
        writer.writerow(["Employee Information"])
        writer.writerow(["Field", "Value"])
        writer.writerow(["Employee ID", employee.employee_id])
        writer.writerow(["Name", employee.name or ""])
        writer.writerow(["Nationality", employee.nationality or ""])
        writer.writerow(["Working Months", employee.working_months or ""])
        writer.writerow(
            [
                "Fetched At",
                (
                    employee.fetched_at.strftime("%Y-%m-%d %H:%M:%S")
                    if employee.fetched_at
                    else ""
                ),
            ]
        )
        writer.writerow([])

        # Write employment details
        if employee.employment_details:
            writer.writerow(["Employment Details"])
            writer.writerow(
                [
                    "Employer",
                    "Status",
                    "Basic Wage",
                    "Housing Allowance",
                    "Other Allowance",
                    "Full Wage",
                ]
            )

            for detail in employee.employment_details:
                writer.writerow(
                    [
                        detail.employer or "",
                        detail.status or "",
                        f"{detail.basic_wage:.2f}" if detail.basic_wage else "0.00",
                        (
                            f"{detail.housing_allowance:.2f}"
                            if detail.housing_allowance
                            else "0.00"
                        ),
                        (
                            f"{detail.other_allowance:.2f}"
                            if detail.other_allowance
                            else "0.00"
                        ),
                        f"{detail.full_wage:.2f}" if detail.full_wage else "0.00",
                    ]
                )

        # Get CSV content and encode as UTF-8 with BOM for Excel compatibility
        csv_content = output.getvalue()
        output.close()

        # Add UTF-8 BOM for Excel to recognize encoding
        csv_bytes = "\ufeff".encode("utf-8") + csv_content.encode("utf-8")

        filename = f"employee_{employee_id}.csv"
        encoded_filename = quote(filename)

        return Response(
            content=csv_bytes,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to export employee CSV: {str(e)}"
        )


@router.get("/database/employee/{employee_id}/excel")
async def export_database_employee_excel(
    employee_id: int,
    db: Session = Depends(deps.get_db),
    current_user: models.User | models.ManagementUser = Depends(
        deps.get_current_active_user_or_management
    ),
) -> Response:
    """
    Export Employee data as Excel (XLSX)
    """
    try:
        from app.models.wathq_employee import Employee
        from sqlalchemy.orm import joinedload
        from urllib.parse import quote
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Fetch employee from database with employment details
        employee = (
            db.query(Employee)
            .options(joinedload(Employee.employment_details))
            .filter(Employee.employee_id == employee_id)
            .first()
        )

        if not employee:
            raise HTTPException(
                status_code=404, detail=f"Employee with ID {employee_id} not found"
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Data"

        # Styling
        header_fill = PatternFill(
            start_color="004074", end_color="004074", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        section_fill = PatternFill(
            start_color="E0ECFF", end_color="E0ECFF", fill_type="solid"
        )
        section_font = Font(bold=True, color="004074")

        # Employee Information Section
        ws["A1"] = "Employee Information"
        ws["A1"].fill = section_fill
        ws["A1"].font = section_font
        ws.merge_cells("A1:B1")

        row = 2
        ws[f"A{row}"] = "Field"
        ws[f"B{row}"] = "Value"
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        ws[f"A{row}"].font = header_font
        ws[f"B{row}"].font = header_font

        row += 1
        ws[f"A{row}"] = "Employee ID"
        ws[f"B{row}"] = employee.employee_id

        row += 1
        ws[f"A{row}"] = "Name"
        ws[f"B{row}"] = employee.name or ""

        row += 1
        ws[f"A{row}"] = "Nationality"
        ws[f"B{row}"] = employee.nationality or ""

        row += 1
        ws[f"A{row}"] = "Working Months"
        ws[f"B{row}"] = employee.working_months or ""

        row += 1
        ws[f"A{row}"] = "Fetched At"
        ws[f"B{row}"] = (
            employee.fetched_at.strftime("%Y-%m-%d %H:%M:%S")
            if employee.fetched_at
            else ""
        )

        # Employment Details Section
        if employee.employment_details:
            row += 2
            ws[f"A{row}"] = "Employment Details"
            ws[f"A{row}"].fill = section_fill
            ws[f"A{row}"].font = section_font
            ws.merge_cells(f"A{row}:F{row}")

            row += 1
            headers = [
                "Employer",
                "Status",
                "Basic Wage",
                "Housing Allowance",
                "Other Allowance",
                "Full Wage",
            ]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for detail in employee.employment_details:
                row += 1
                ws[f"A{row}"] = detail.employer or ""
                ws[f"B{row}"] = detail.status or ""
                ws[f"C{row}"] = float(detail.basic_wage) if detail.basic_wage else 0.0
                ws[f"D{row}"] = (
                    float(detail.housing_allowance) if detail.housing_allowance else 0.0
                )
                ws[f"E{row}"] = (
                    float(detail.other_allowance) if detail.other_allowance else 0.0
                )
                ws[f"F{row}"] = float(detail.full_wage) if detail.full_wage else 0.0

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 15

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"employee_{employee_id}.xlsx"
        encoded_filename = quote(filename)

        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to export employee Excel: {str(e)}"
        )
